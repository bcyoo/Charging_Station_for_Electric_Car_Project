# -*- coding: utf-8 -*-
# ---
# jupyter:
#   jupytext:
#     formats: ipynb,py:light
#     text_representation:
#       extension: .py
#       format_name: light
#       format_version: '1.5'
#       jupytext_version: 1.13.1
#   kernelspec:
#     display_name: Python 3 (ipykernel)
#     language: python
#     name: python3
# ---

# +
# 한글폰트 사용
import platform
from matplotlib import font_manager, rc
plt.rcParams['axes.unicode_minus'] = False

if platform.system() == 'Darwin':
    f_path = '/Library/Fonts/Arial Unicode.ttf'
elif platform.system() == 'Windows':
    f_path = 'c:/Windows/Fonts/malgun.ttf'
font_name = font_manager.FontProperties(fname=f_path).get_name()
rc('font', family=font_name)

print('Hangul font is set!')

# +
# 필요패키지 import
import numpy as np
import pandas as pd
from datetime import datetime
import csv # csv 파일 저장
import matplotlib.pyplot as plt
import seaborn as sns
from plotnine import * # 데이터 시각화 
import re # 정규표현식
# 지도 표현
import folium

# %matplotlib inline 
# -

## geopy로 주소 간단하게 지오코딩 하는 법
from geopy.geocoders import Nominatim
import osmnx as ox 

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)

df = pd.read_csv('DATA/시군구_자동차_등록현황_2021_10.csv', engine='python')
df

df["시군구"] = df["시"].map(str) + " " + df["군구"]

df = df.loc[df['군구'] != '계']
df

df = df1.loc[df['시'] != '합계' ]
df

df.columns.tolist()

df = df[['시군구', '시', '군구', '자동차 등록 현황']]
df

df.to_excel(excel_writer='address.xlsx',encoding = 'cp949')

# +
# key : E07005C0-0CEE-30AC-B9D7-84F23F520E05
#-*-coding: utf-8 -*-
import requests
import openpyxl
import json

## 엑셀 파일 오픈
filename = "address.xlsx"
exelFile = openpyxl.load_workbook(filename)

## 시트 설정
sheet = exelFile.worksheets[0]

## 데이터 가져오기
rowCount = 1
for row in sheet.rows:	

	try:
		## geocoder 호출
		r = requests.get("http://apis.vworld.kr/new2coord.do?q="+row[0].value+"&output=json&epsg=epsg:4326&apiKey=E07005C0-0CEE-30AC-B9D7-84F23F520E05")

		## cell 설정 [ (10)J1 ~ J* : 위도 / (11)K1 ~ K* : 경도]
		lat_cell = sheet.cell(row = rowCount, column = 2)
		lng_cell = sheet.cell(row = rowCount, column = 3)

		## 데이터 추가
		lat_cell.value = r.json()["EPSG_4326_Y"]
		lng_cell.value = r.json()["EPSG_4326_X"]
	except KeyError as ke:
		lat_cell.value = 0
		lng_cell.value = 0
	except TypeError as te:
		lat_cell.value = 0
		lng_cell.value = 0
		
	rowCount = rowCount + 1

## 데이터 저장
exelFile.save("address.xlsx")
# -

df_1 = pd.read_csv('./address.csv')
df_1

(ggplot(df_1)  ## gg plot 으로 시각화
 + aes(x='경도', y='위도', color='자동차 등록 현황')
 + geom_point() # 점포인트로 찍음
 + theme(text=element_text(family=font_name))
) 

df_1

df

df_1.drop(columns=['시군구', '자동차 등록 현황'], inplace=True)
df_1

df_1

df_2 = df.merge(df_1, left_index=True, right_index=True)
df_2.info()

df_2.columns = ['주소', '시도', '구군', '자동차등록현황', '위도', '경도']
df_2

# +
df_2.shape
map = folium.Map(location=[df_2['위도'].mean(), df_2['경도'].mean()], zoom_start=13)

for n in df_2.index:
    park_name = df_2.loc[n, '자동차등록현황'] + ' - ' + df_2.loc[n, '주소']
    folium.Marker([df_2.loc[n, '위도'], df_2.loc[n, '경도']], popup=park_name).add_to(map)
map
